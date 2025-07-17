import requests
import json
import pandas as pd
from requests.auth import HTTPBasicAuth
import openpyxl
import html 
import re
import os
import truststore
from templates import (
    build_description_html,
    build_grouped_description_html,
    build_acceptance_criteria_html,
    render_grouped_remediations,
    render_single_remediation,
    build_custom_ac_list,
    build_custom_ac_paragraph,
    build_grouped_acceptance_criteria_html
)
from resource_lookup import build_resource_lookup


truststore.inject_into_ssl()

from dotenv import find_dotenv, load_dotenv

dotenv_path = find_dotenv()
load_dotenv(dotenv_path)

# TFS configuration
ORG_URL = os.getenv("ORG_URL")
PROJECT = "Design"  # The project you're targeting
API_VERSION = "6.0"  # Update to a newer version for better support

# Function to create a PBI
def create_pbi(title, description, acceptance_criteria, priority, tags, pat):
    url = f"{ORG_URL}/{PROJECT}/_apis/wit/workitems/$Product%20Backlog%20Item?api-version={API_VERSION}"

    headers = {
        "Content-Type": "application/json-patch+json"
    }

    # Body for the work item creation
    body = [
        {"op": "add", "path": "/fields/System.Title", "value": title},
        {"op": "add", "path": "/fields/System.Description", "value": description},
        {"op": "add", "path": "/fields/Microsoft.VSTS.Common.AcceptanceCriteria", "value": acceptance_criteria},
        {"op": "add", "path": "/fields/Microsoft.VSTS.Common.Priority", "value": priority},
        {"op": "add", "path": "/fields/System.IterationPath", "value": "Design\\Accessibility"},
        {"op": "add", "path": "/fields/System.AreaPath", "value": "Design\\Accessibility"},
        {"op": "add", "path": "/fields/System.Tags", "value": tags}
    ]

    try:
        # Send the request to create the work item
        response = requests.post(url, headers=headers, data=json.dumps(body), auth=HTTPBasicAuth('', pat))

        if response.status_code in (200, 201):
            print(f"Successfully created PBI: {response.json()['id']} at {response.json()['_links']['html']['href']}\n")
            return response.json()['id']  # Return the ID of the created PBI
        else:
            print(f"Failed to create PBI for {title}. Status Code: {response.status_code}, Response: {response.text}\n")
            return None
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# Function to link the created PBI to a Parent Feature
def link_pbi_to_feature(pbi_id, feature_id, pat):
    url = f"{ORG_URL}/{PROJECT}/_apis/wit/workitems/{pbi_id}?api-version={API_VERSION}"

    # Define the relation to link the PBI to the Parent Feature
    relation_body = [
        {
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": f"{ORG_URL}/_apis/wit/workItems/{feature_id}",
                "attributes": {"comment": "Linking PBI to Parent Feature"}
            }
        }
    ]

    response = requests.patch(url, headers={"Content-Type": "application/json-patch+json"}, data=json.dumps(relation_body), auth=HTTPBasicAuth('', pat))

    if response.status_code in (200, 204):
        pass
    else:
        print(f"ERROR: Failed to link PBI. Status Code: {response.status_code}, Response: {response.text}")

# Function to map priority from text to numerical value
def map_priority(priority_text):
    priority_map = {"High": 1, "Medium": 2, "Low": 3}
    return priority_map.get(priority_text, 3)  # Default to Low (3) if not found

def write_pbi_url_to_excel(workbook, summary_sheet, row_index, pbi_url):
    # Find the index of the "Remediation PBI" column
    remediation_pbi_column = None
    
    # Iterate through the columns in the first row (headers) to find the "Remediation PBI" column
    for col in range(1, summary_sheet.max_column + 1):
        cell_value = summary_sheet.cell(row=1, column=col).value
        if cell_value == "Remediation PBI":
            remediation_pbi_column = col
            break
    
    # Write the PBI URL as a clickable hyperlink
    if remediation_pbi_column:
        cell = summary_sheet.cell(row=row_index, column=remediation_pbi_column)
        cell.value = pbi_url
        cell.hyperlink = pbi_url
        cell.style = "Hyperlink"
    else:
        print("ERROR: 'Remediation PBI' column not found in the sheet.")


# Helper to safely convert a value to string if needed
# and escape HTML characters to prevent injection 
def safe_html(val):
    return html.escape(str(val)) if pd.notna(val) else ""

def format_custom_ac(raw_text):
    # Converts raw Acceptance Criteria text into styled HTML.
    # If the text appears to contain numbered items like '1. Step one 2. Step two',
    # it will split them into an unordered <ul> list.
    # Otherwise, it will wrap the text in a <p>.
    
    # Split on patterns like '1. ', '2. ', etc.
    potential_items = re.split(r'\d+\.\s*', raw_text)
    cleaned_items = [item.strip() for item in potential_items if item.strip()]

    # Decide whether to render as list or paragraph
    if len(cleaned_items) > 1:
        # Build list items
        list_items_html = "".join(
            f"<li>{html.escape(item)}</li>" for item in cleaned_items
        )
        return build_custom_ac_list(list_items_html)
    else:
        # Just a single paragraph
        return build_custom_ac_paragraph(html.escape(raw_text))

# Main function to read the Excel file and create PBIs
def create_pbis_from_excel(excel_path, pat):
    try:
        # Open the workbook and sheets
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        report_details_sheet = workbook['Report Details']
        summary_sheet = workbook['Evaluation']
        resource_lookup = build_resource_lookup(workbook)
        
        # Extract information from the 'Report Details' sheet
        page_name = report_details_sheet.cell(row=6, column=2).value
        page_url = report_details_sheet.cell(row=4, column=2).value
        feature_id = report_details_sheet.cell(row=12, column=2).value
        testing_account_cell = report_details_sheet.cell(row=5, column=2)

        #ensure feature_id is not empty
        if not feature_id:
            print("ERROR: Feature ID is missing from the Report Details sheet. This must be filled in before creating PBIs.")
            print("Exiting script early — no PBIs were created.\n")
            return

        # Check that the page URL points to the development environment
        if str(page_url).startswith("https://www.webstaurantstore.com"):
            # If not a development URL, replace with the development URL
            page_url_base = "https://www.dev.webstaurantstore.com"
            page_url = page_url.replace("https://www.webstaurantstore.com", page_url_base)
        
        # Check that feature_id is not a hyperlink
        if feature_id and str(feature_id).startswith("https://"):
            # If it's a hyperlink, extract the ID from the URL
            if "?" in feature_id:
                feature_id = feature_id.split("=")[-1]
            else:
                feature_id = feature_id.split("/")[-1]
        
        # Check if the testing account cell has a hyperlink (indicating it's valid)
        if testing_account_cell.hyperlink:
            testing_account_url = testing_account_cell.hyperlink.target
            testing_account_html = f'<ul><li>Log in with <a href="{html.escape(testing_account_url)}">this account</a></li></ul>'
        else:
            testing_account_html = ""  # If no hyperlink, leave it empty

        # Read the 'Evaluation' sheet into a DataFrame
        summary = pd.read_excel(excel_path, sheet_name='Evaluation', engine='openpyxl')

        # Get the column index of the 'Resources, Screen Captures, Links' column
        resources_column_index = summary.columns.get_loc('Resources, Screen Captures, Links') + 1  # +1 for openpyxl index
    
        # 1) Load the whole DataLayer sheet into memory
        data_layer_sheet = pd.read_excel(
            excel_path,
            sheet_name="DataLayer",
            engine="openpyxl",
            dtype=str
        )
        # 2) Clean up the Acceptance Criteria column and drop truly empty rows
        data_layer_sheet["Acceptance Criteria"] = (
            data_layer_sheet["Acceptance Criteria"]
            .fillna("")       # replace NaN with empty string
            .str.strip()      # trim whitespace
        )
        # Keep only rows that actually have AC text
        rows_with_custom_ac = data_layer_sheet[
            data_layer_sheet["Acceptance Criteria"] != ""
        ]

        # 3) Build a lookup so we can quickly find AC by (Notes, Remediation) key
        acceptance_criteria_lookup: dict[tuple[str,str], str] = {}
        for _, row in rows_with_custom_ac.iterrows():
            notes_key       = row["Notes"].strip()
            remediation_key = row["Remediation Techniques"].strip()
            ac_text         = row["Acceptance Criteria"]
            acceptance_criteria_lookup[(notes_key, remediation_key)] = ac_text

        # Pre-aggregate remediation techniques for grouped rows
        grouped_data = {}
        if "Group" in summary.columns:
            for index, row in summary.iterrows():
                group_val = row.get("Group")
                if pd.notna(group_val):
                    excel_row_number = index + 2  
                                     
                    # Build the list of resource entries for grouped PBIs.
                    resource_entries = []
                    for col_index in range(resources_column_index, summary_sheet.max_column + 1):
                        cell = summary_sheet.cell(row=excel_row_number, column=col_index)
                        resource_text = cell.value.strip() if cell.value else None
                        if resource_text:
                            if cell.hyperlink:
                                url = cell.hyperlink.target
                                resource_entries.append(f'<li><a href="{safe_html(url)}">{safe_html(resource_text)}</a></li>')
                            elif resource_text in resource_lookup:
                                url = resource_lookup[resource_text]
                                resource_entries.append(f'<li><a href="{safe_html(url)}">{safe_html(resource_text)}</a></li>')
                            else:
                                resource_entries.append(f'<li>{safe_html(resource_text)}</li>')

                    # Look up custom AC for this row
                    ac_text = acceptance_criteria_lookup.get(
                        (str(row.get("Notes", "")).strip(), str(row.get("Remediation Techniques", "")).strip())
                    )
                    if group_val not in grouped_data:
                        grouped_data[group_val] = []

                     # Append entry with AC included    
                    grouped_data[group_val].append({
                        "recommendation": safe_html(row.get("Conformance Recommendation", "")),
                        "notes": safe_html(row.get("Notes", "")),
                        "remediation": safe_html(row.get("Remediation Techniques", "")),
                        "description": safe_html(row.get("Description", "")),
                        "resources": resource_entries,
                        "ac": ac_text
                    })


        
        # Dictionary to store created PBI URL for each group
        group_pbi_map = {}

        # List to hold PBI URLs for writing after PBIs are created
        pbi_urls = []

        # Loop through each row in the DataFrame and create PBIs
        for index, row in summary.iterrows():
            # Skip rows that have a value in the 'Remediation PBI' column
            if pd.notna(row.get('Remediation PBI')):
                print(f"Skipped row {index + 2}, PBI already assigned.\n")
                continue

            # Skip rows that are Compliant
            if str(row.get('Conformance', '')).strip().lower() != "non-compliant":
                print(f"Skipped row {index + 2}\n")
                continue

            # Determine if this row is part of a group
            group_val = row.get("Group")

            # Map the columns to the corresponding PBI fields
            title = f"Remediation - {page_name} - {row['Notes'][:50]}"  # Limiting title to 50 characters

            # Escape the content to prevent HTML injection
            page_name_escaped = html.escape(page_name)
            page_url_escaped = html.escape(page_url)
            notes_escaped = safe_html(row.get('Notes', ''))
            recommendation_escaped = safe_html(row.get('Conformance Recommendation', ''))
            remediation_escaped = safe_html(row.get('Remediation Techniques', ''))
            description_escaped = safe_html(row.get('Description', ''))
           
            # Check if this row is part of a group and render the remediation list accordingly           
            if pd.notna(group_val):
                remediation_list = render_grouped_remediations(grouped_data.get(group_val, []))
            else:
                remediation_list = render_single_remediation(remediation_escaped, description_escaped)
            
            # Build the resources list, only adding non-empty cells (ignoring whether there's a hyperlink or not)
            resources_list = []

            # Calculate the Excel row number (pandas uses 0-based, Excel uses 1-based)
            excel_row_number = index + 2

            # Get the 'Resources' cell from openpyxl
            resource_cell = summary_sheet.cell(row=excel_row_number, column=resources_column_index)
            resource_text = resource_cell.value.strip() if resource_cell.value else None

            if resource_text:
                # Case 1: Manually inserted hyperlink
                if resource_cell.hyperlink:
                    url = resource_cell.hyperlink.target
                    resources_list.append(f'<li><a href="{safe_html(url)}">{safe_html(resource_text)}</a></li>')
                    print(f"Using manual hyperlink for '{resource_text}': {url}")

                # Case 2: Lookup in DataLayer sheet
                elif resource_text in resource_lookup:
                    url = resource_lookup[resource_text]
                    resources_list.append(f'<li><a href="{safe_html(url)}">{safe_html(resource_text)}</a></li>')
                    print(f"Resolved '{resource_text}' from DataLayer to: {url}")

                # Case 3: Just text, fallback
                else:
                    resources_list.append(f'<li>{safe_html(resource_text)}</li>')
                    print(f"No link found for '{resource_text}', using plain text")


            # Now check the columns beyond 'Resources' (starting from the next column)
            for col_index in range(resources_column_index + 1, summary_sheet.max_column + 1):
                resource_cell = summary_sheet.cell(row=excel_row_number, column=col_index)

                # Check if the cell has a value
                if resource_cell.value:
                    resource_content = safe_html(resource_cell.value)
                    # Check if the cell has a hyperlink; if not, use the cell's value_cell.value
                    hyperlink = resource_cell.hyperlink.target if resource_cell.hyperlink else None
                    if hyperlink is not None:
                        resources_list.append(f'<li><a href="{hyperlink}">{resource_content}</a></li>')
                    else:
                        resources_list.append(f'<li>{resource_content}</li>')

            # Only generate the <ul> block if there's actual resource content
            resources_html = "".join(resources_list) if resources_list else ""
       
            # Build the description HTML based on whether it's a grouped row or not
            if pd.notna(group_val):
                description = build_grouped_description_html(
                        page_name_escaped,
                        page_url_escaped,
                        testing_account_html,
                        remediation_list
                )
            else:
                description = build_description_html(
                    page_name_escaped,
                    page_url_escaped,
                    testing_account_html,
                    recommendation_escaped,
                    notes_escaped,
                    remediation_list,
                    resources_html
        )
         
            # CUSTOM vs DEFAULT Acceptance Criteria
            if pd.notna(group_val):
                # For grouped PBIs, build an ordered list of all ACs
                acceptance_criteria = build_grouped_acceptance_criteria_html(
                    grouped_data.get(group_val, []),
                    format_custom_ac
                )
            else:
                # For non-grouped PBIs, use single-item logic
                note = str(row.get("Notes", "")).strip()
                rem  = str(row.get("Remediation Techniques", "")).strip()
                raw_ac = acceptance_criteria_lookup.get((note, rem))
                if raw_ac:
                    acceptance_criteria = format_custom_ac(raw_ac)
                else:
                    acceptance_criteria = build_acceptance_criteria_html(
                        page_url_escaped,
                        page_name_escaped
                    )


            priority = map_priority(row['Priority'])
            tags = f"Remediation,Accessibility,{page_name} Page"

            # Process grouped rows differently
            if pd.notna(group_val):
                if group_val in group_pbi_map:
                    # Already created a PBI for this group; reuse its URL
                    pbi_url = group_pbi_map[group_val]
                    print(f"Using existing PBI for group {group_val} at row {index+2}")
                else:
                    print(f"Creating grouped PBI for group {group_val} at row {index+2}...")
                    pbi_id = create_pbi(title, description, acceptance_criteria, priority, tags, pat)
                    if pbi_id:
                        pbi_url = f"{ORG_URL}/_workitems/edit/{pbi_id}"
                        group_pbi_map[group_val] = pbi_url  # Store for later reuse
                        # Link to parent feature (only once per group)
                        link_pbi_to_feature(pbi_id, feature_id, pat)
                    else:
                        pbi_url = None
                if pbi_url:
                    pbi_urls.append((index + 2, pbi_url))
                else:
                    print(f"ERROR: Failed to create PBI for grouped row {index+2}.")
            else:
                # Process non-grouped row normally
                pbi_id = create_pbi(title, description, acceptance_criteria, priority, tags, pat)
                if pbi_id:
                    pbi_url = f"{ORG_URL}/_workitems/edit/{pbi_id}"
                    pbi_urls.append((index + 2, pbi_url))
                    link_pbi_to_feature(pbi_id, feature_id, pat)

        # Now that all PBIs are created, write PBI URLs to the Excel sheet
        for row_index, pbi_url in pbi_urls:
            write_pbi_url_to_excel(workbook, summary_sheet, row_index, pbi_url)

        print("\nUPDATED: All PBI URLs written into Excel file\n")

        # Save the workbook after writing all URLs
        workbook.save(excel_path)
        
        print("\nSUCCESS: PBI creation complete!")
    
    except FileNotFoundError:
        print(f"ERROR: File {excel_path} not found. Please check the path and try again.")
    except Exception as e:
        print(f"ERROR: An error occurred: {str(e)}")


if __name__ == "__main__":
    # Get Personal Access Token from user
    PAT = input("Please enter your ADO Personal Access Token (PAT): ")

    attempts = 0  # Initialize attempt counter
    max_attempts = 3  # Set maximum number of attempts

    while attempts < max_attempts:  # Limit attempts to max_attempts
        # Prompt user for the Excel file path
        excel_file_path = input("Please enter the path to the Excel file: ")

        # Check if the file exists
        if os.path.isfile(excel_file_path):
            # Create PBIs from the Excel file
            create_pbis_from_excel(excel_file_path, PAT)
            break  # Exit the loop if the file is valid
        else:
            attempts += 1  # Increment the attempt counter
            print(f"ERROR: The file '{excel_file_path}' does not exist. Please provide a valid path.")

            # Provide feedback on remaining attempts
            remaining_attempts = max_attempts - attempts
            if remaining_attempts > 0:
                print(f"You have {remaining_attempts} attempt(s) left.")
            else:
                print("You have exceeded the maximum number of attempts. Exiting the program.")
