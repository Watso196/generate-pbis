import html
import re
import pandas as pd
from openpyxl import load_workbook

def safe_html(val):
    # Safely convert a value to string if needed and escape HTML characters to prevent injection.
    return html.escape(str(val)) if pd.notna(val) else ""


import re
import html

def format_custom_acceptance_criteria(raw_text, page_url, testing_account_html):
    # Converts raw Acceptance Criteria text into styled HTML.
    # Splits on '1. ', '2. ', etc. for main items.
    # Then inside each main item, splits on '*' optionally for sub‑bullets.
    # Otherwise, wraps everything in a <p>.

    # Do we have at least one top‑level numbered bullet?
    is_list_mode = bool(re.search(r'(?m)^\d+\.\s*', raw_text))

    # 1) Split out each numbered bullet (anchored at line start)
    potential_items = re.split(r'(?m)^\d+\.\s*', raw_text)
    cleaned_items = [item.strip() for item in potential_items if item.strip()]

    if is_list_mode:
        list_items_html = ""

        for main_item in cleaned_items:
            # 2) Split off any sub‑bullets (asterisks at line start)
            asterisk_items = re.split(r'(?m)^\*\s*', main_item)
            main_line = asterisk_items[0].strip()
            sub_items = [
                sub.strip() for sub in asterisk_items[1:]
                if sub.strip()
            ]

            if sub_items:
                # Build nested UL
                sub_html = "".join(f"<li>{html.escape(sub)}</li>"
                                   for sub in sub_items)
                list_items_html += (
                    f"<li>{html.escape(main_line)}"
                    f"<ul>{sub_html}</ul>"
                    "</li>"
                )
            else:
                # No sub‑bullets, just render the main line
                list_items_html += f"<li>{html.escape(main_line)}</li>"

        from templates import build_custom_acceptance_criteria_list
        # Pass page_url and testing_account_html so the builder can prepend the visit link
        return build_custom_acceptance_criteria_list(
            list_items_html,
            page_url,
            testing_account_html
        )
    else:
        # Single item -> fallback to paragraph (in case the AC convention is broken)
        from templates import build_custom_acceptance_criteria_paragraph
        # Pass page_url and testing_account_html so the builder can prepend the visit link
        return build_custom_acceptance_criteria_paragraph(
            html.escape(raw_text),
            page_url,
            testing_account_html
        )

def build_resource_lookup(workbook):
    #Extracts a resource lookup from the DataLayer to provide a hyperlinked entry for the PBIs.
    sheet = workbook['DataLayer']
    resource_lookup = {}

    for row in range(2, sheet.max_row + 1):
        friendly_text = sheet.cell(row=row, column=5).value  # Column E
        url = sheet.cell(row=row, column=6).value            # Column F

        if friendly_text and url:
            resource_lookup[friendly_text.strip()] = url.strip()

    return resource_lookup
