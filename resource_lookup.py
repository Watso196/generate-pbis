from openpyxl import load_workbook

# This function extracts a resource lookup from the DataLayer 
# to provide a hyperlinked entry for the PBIs 
def build_resource_lookup(workbook):
    sheet = workbook['DataLayer']
    resource_lookup = {}

    for row in range(2, sheet.max_row + 1):
        friendly_text = sheet.cell(row=row, column=5).value  # Column E
        url = sheet.cell(row=row, column=6).value            # Column F

        if friendly_text and url:
            resource_lookup[friendly_text.strip()] = url.strip()

    return resource_lookup
